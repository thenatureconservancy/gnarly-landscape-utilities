# ---------------------------------------------------------------------------
# HabitatAndResistanceCalculator.py
# Based on code written by Andrew Shirk, University of Washington, CSES Climate Impacts Group
# Modified by Brad McRae and Jim Platt, TNC
# 
# Required Software:
# ArcGIS 10.x with Spatial Analyst extension
# Python 2.6 or 2.7
# ---------------------------------------------------------------------------


__filename__ = "HabitatAndResistanceCalculator.py"  

# Import system modules
import sys
import os
import time
import traceback  
import shutil
import numpy as npy
import gnarly_version as ver

__version__ = ver.releaseNum

from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook

import arcgisscripting
import arcpy

arcpy.CheckOutExtension("Spatial")
gp = arcpy.gp      
arcpy.env.overwriteOutput = True

projectFolder = sys.argv[3]       

def create_dir(lmfolder):
    """Creates folder if it doesn't exist."""
    if not os.path.exists(lmfolder):
        os.makedirs(lmfolder)    

file,ext=os.path.splitext(projectFolder)
if ext == '.gdb':
    arcpy.AddError('Error: output directory must be a folder, not a geodatabase.')
    # for msg in range(0, gp.MessageCount):
        # if gp.GetSeverity(msg) == 2:
            # gp.AddReturnMessage(msg)
            
    if not arcpy.GetMessages(2) == "":
        arcpy.AddError(arcpy.GetMessages(2))                                
            
    exit(1)
messageDir = os.path.join(projectFolder,'messages')
scratchDir = os.path.join(projectFolder,'scratch')


doExpandCells = True # Temporarily disable this feature

create_dir(messageDir)
create_dir(scratchDir)
        
ft = tuple(time.localtime())
timeNow = time.ctime()
fileName = ('%s_%s_%s_%s%s_%s.txt' % (ft[0], ft[1], ft[2], ft[3], ft[4], 'H_R_Calc'))
logFilePath = os.path.join(messageDir,fileName)

try:
    logFile=open(logFilePath,'a')
except:
    logFile=open(logFilePath,'w')

if sys.argv is not None:
    logFile.write('*'*70 + '\n')
    logFile.write('Habitat and Resistance Calculator log file: %s \n\n' % ('HR'))
    logFile.write('Start time:\t%s \n' % (timeNow))
    logFile.write('Parameters:\t%s \n\n' % (sys.argv))
logFile.close()
            
def habitat_model_builder():
    try: 
        check_path(projectFolder)
        # Local variables...
        GP_NULL = '#'
        tables = sys.argv[1]
        if " " in tables:
            arcpy.AddError('Error: spaces are not allowed in spreadsheet file names.')
            # for msg in range(0, gp.MessageCount):
                # if gp.GetSeverity(msg) == 2:
                    # gp.AddReturnMessage(msg)
                    
            if not arcpy.GetMessages(2) == "":
                arcpy.AddError(arcpy.GetMessages(2))                                
                    
            exit(1)
        
        gprint('\nHabitat and Resistance Calculator version ' + __version__)
        gprint('\n-----------------------------------------------------------------')
        gprint('If you use this software, please cite it so others can find it!')
        gprint('See www.circuitscape.org/gnarly-landscape-utilities \nfor preferred citation')
        gprint('-----------------------------------------------------------------')

        gprint('\nProcessing the following Excel parameter tables:\n%s' %tables)
        tables = tables.split(';')
        layerFolder = sys.argv[2] 
        
        projectfolder = sys.argv[3] + '\\'
        habitatMethod = sys.argv[4] 
        if habitatMethod == GP_NULL:
            habitatMethod = 'NONE'
        resistMethod = sys.argv[5]
        if resistMethod == GP_NULL:
            resistMethod = 'NONE'
        addOneToResist = str2bool(sys.argv[6])

        classIDColumn = 'B'
        LIVariableColumn = 'E'
        resistanceVariableColumn = 'F'
        expandCellsColumn = 'G'
        
        gprint('\nLayer name should be in spreadsheet column A')
        gprint('Class ID should be in column %s' %classIDColumn)
        gprint('Habitat quality/suitability scores should be in column %s' %LIVariableColumn)
        gprint('Resistance values should be in column %s' %resistanceVariableColumn)
        gprint('EXTENT and CELL SIZE of output will be based on the first layer.\n')
        
        if (habitatMethod == 'NONE' or habitatMethod == "'NONE'") and (resistMethod == 'NONE' or resistMethod == "'NONE'"):
            gprint('*' * 45)
            gprint('Both habitat and resistance methods are set to none! Bailing.')
            gprint('*' * 45)
            gprint('\n')
            return
        
        for iteration in range(1,3):
            if iteration == 1:
                if resistMethod == 'NONE' or resistMethod == "'NONE'":
                    continue
                task = 'RESISTANCE'
                gprint('*' * 45)
                gprint('Calculating Resistance Values')
                method = resistMethod   
                variableColumn = resistanceVariableColumn
            else:
                if habitatMethod == 'NONE' or habitatMethod == "'NONE'":
                    continue
                task = 'HABITAT'
                gprint('*' * 45)
                gprint('Calculating habitat Values')
                method = habitatMethod
                variableColumn = LIVariableColumn
                
            gprint('Using %s method on Excel column %s' %(method, variableColumn))
            # Iterate habitat model for each table in table list
            for tableName in tables:
                tableBase,ext = os.path.splitext(os.path.basename(tableName))
                # species = species_tbl[0:-5]
                create_dir(projectfolder)
                    
                outputGDB = os.path.join(projectfolder,'habitat_resis_layers.gdb')
                # delete_data(outputGDB) #Caused file locks
                if not arcpy.Exists(outputGDB):
                    arcpy.CreateFileGDB_management(projectfolder, os.path.basename(outputGDB))
                try:
                    arcpy.Compact_management(outputGDB)
                except:
                    if arcpy.Exists(outputGDB):
                        shutil.rmtree(outputGDB)
                    if not arcpy.Exists(outputGDB):
                        gprint('Recreating output GDB')
                        arcpy.CreateFileGDB_management(projectfolder, os.path.basename(outputGDB))
                    # clearWSLocks(outputGDB)
                          
                gprint('\n%s\n' %('*' * 45))
                gprint('READING TABLE: %s\n' %tableBase)

                # open and read input file
                wb = load_workbook(filename=tableName)
                ws = wb.get_active_sheet()
                # get datalayer names and ranges of parameters in table
                datarange = 'A2' + ':A' + str(len(ws.row_dimensions))
                layer_range = ws.range(datarange)
                layerlist = []  # a new list

                # number of 'rows' imported to matm tuple
                row = range(len(layer_range))
                # number of 'columns' imported to matm tuple
                column = range(len(layer_range[0]))

                for i in row:
                    #traverse columns while keeping row fixed
                    nvalues = [layer_range[i][j].value for j in column]
                    layerlist.append(str(nvalues)[3:(len(str(nvalues)) - 2)])
                    
                    # 1st layer will be used to for environment settings
                    if i == 0:
                        referenceLayer = str(nvalues)[3:(len(str(nvalues)) - 2)]
                        gprint("******************************************************")
                        gprint("Extent and cell size of outputs will be based on first")
                        gprint("input layer in spreadsheet: " + referenceLayer)                    
                        gprint("******************************************************\n")
                        
                layers = unique(layerlist)
                gprint('Input Layers: %s' %str(layers))
                for layer in layers:
                    if layer == 'n':
                        continue
                    rows = []
                    for i in range(0, len(layerlist), 1):
                        if layer == layerlist[i]:
                            rows.append(i + 1)
                    vars()[layer + '_range'] = range(min(rows) + 1, len(rows) + 1 + min(rows), 1)
                    try:
                        uniqueValueCt = arcpy.GetRasterProperties_management(os.path.join(layerFolder,layer),"UNIQUEVALUECOUNT")
                    except:    
                        arcpy.BuildRasterAttributeTable_management(os.path.join(layerFolder,layer), "Overwrite")
                        uniqueValueCt = arcpy.GetRasterProperties_management(os.path.join(layerFolder,layer),"UNIQUEVALUECOUNT")                    
                    if len(rows)>int(uniqueValueCt.getOutput(0)):
                        #xxx BHM 11/7/13- may not need this error?
                        arcpy.AddWarning('Warning: there is at least value in the excel spreadsheet for layer "' + layer + '"\n'
                        'that has no corresponding value in the layer.  This may cause reclass problems.') 
                        
                        # arcpy.AddError('Error: there is at least one entry in the excel spreadsheet that has no') #Needed because we rely on count.
                        # arcpy.AddError('corresponding value in the raster layer "' + layer + '".')
                        # # for msg in range(0, gp.MessageCount):
                            # # if gp.GetSeverity(msg) == 2:
                                # # gp.AddReturnMessage(msg)
                                
                        # if not arcpy.GetMessages(2) == "":
                            # arcpy.AddError(arcpy.GetMessages(2))                                
                                
                        # exit(1)
                    if len(rows)<int(uniqueValueCt.getOutput(0)):
                        arcpy.AddWarning('Warning: there is at least one raster cell value in layer "' + layer + '"')
                        arcpy.AddWarning('that has no corresponding value in the excel spreadsheet.  Cells with this') 
                        arcpy.AddWarning('value will be left out of calculations.\n')

                #Generate remap table and write to outputGDB
                gprint('GENERATING REMAP TABLES\n')
                for layer in layers:
                    if layer == 'n':
                        continue
                    remapFile = open(os.path.join(scratchDir, layer + '_' + task + '_remap.txt'), 'w')
                    # remapFile.write('From')
                    # remapFile.write('\t')
                    # remapFile.write('To')
                    # remapFile.write('\n')
                    
                    # expandTable = []
                    # expandValueList = []
                    
                    
                    for category in range(1, len(vars()[layer + '_range']) + 1):
                        cell = variableColumn + str(vars()[layer + '_range'][category - 1])
                        classID = classIDColumn + str(vars()[layer + '_range'][category - 1])
                        
                        ### multiply by 1000 because reclassify uses integers
                        classID_value = ws.cell(classID).value
                        value = int(float(ws.cell(cell).value) * 1000) 
                        if category == 1:
                            values = npy.array([[classID_value,value]])
                        else:
                            values = npy.append(values, [[classID_value,value]], axis=0)
                    if len(values) > 0: #Sort in ascending order
                        ind = npy.argsort((values[:, 0]))  # Sort by classID_value
                        values = values[ind]    
                    
                    for category in range(0, len(vars()[layer + '_range'])):
                        remapFile.write(str(values[category,0]))
                        remapFile.write(' : ')
                        remapFile.write(str(values[category,1]))
                        remapFile.write('\n')

                    vars()[layer + '_values'] = values
                    remapFile.close()
                #Determine which layers are to be used
                usedlayers = []

                for layer in layers:
                    #if min(vars()[layer + '_values']) < 1000:  
                    if layer != 'n':
                        usedlayers.append(layer)

                gprint('THE FOLLOWING LAYERS WILL BE INCLUDED IN CALCULATIONS:')
                for layer in usedlayers:
                    layerRow = str(vars()[layer + '_range'][0])
                    expandCellsValue = ws.cell(expandCellsColumn + layerRow).value
                    if expandCellsValue > 0 and task == 'RESISTANCE' and doExpandCells == True:
                        gprint('\t%s' %layer)
                        if method == 'MINIMUM' or method == "'MINIMUM'": 
                            gprint('    ***MINIMUM value for reclassified layer ' + layer + ' will be expanded by ' + str(expandCellsValue) + ' cell(s).')
                        else:
                            gprint('  ***Maximum value for ' + layer + 'layer will be expanded by ' + str(expandCellsValue) + ' cell(s)'
                               ' for resistance calculations.')
                    else:
                        gprint('\t%s' %layer)

                #Reclassify layer by remap table
                arcpy.env.overwriteOutput = True
                arcpy.env.workspace = layerFolder
                scratchGDB = os.path.join(scratchDir,'scratchGDB'+task+'.gdb')
                # clean_out_ws(scratchGDB)
                delete_data(scratchGDB)
                delete_dir(scratchGDB)
                if not arcpy.Exists(scratchGDB):
                    arcpy.CreateFileGDB_management(scratchDir, 'scratchGDB'+task+'.gdb')
                counter = 0
                for layer in usedlayers:
                    gprint('Reclassifying ' + layer)
                    counter += 1
                    layerPath = os.path.join(layerFolder,layer)
                    # uniqueValueCt = arcpy.GetRasterProperties_management(layerPath,"UNIQUEVALUECOUNT")
                    remapFile = os.path.join(scratchDir, str(layer)  + '_' + task + '_remap.txt')
                    arcpy.env.workspace = scratchGDB
                    arcpy.env.scratchWorkspace = scratchGDB
                    try:               
                        outReclass = arcpy.sa.ReclassByASCIIFile(layerPath, remapFile) # Reclass values can't be too large?
                    except:
                        arcpy.AddError('Reclass failed.  There may be an entry in the excel spreadsheet that has no')
                        arcpy.AddError('corresponding value in the raster being reclassified, or classes may not be')
                        arcpy.AddError('in ascending order, or habitat/resistance value may be > 1000000 maximum.')
                        arcpy.AddError('Re-starting ArcGIS or selecting a new output directory sometimes fixes this error.\n')
                        if not arcpy.GetMessages(2) == "":
                            arcpy.AddError(arcpy.GetMessages(2)) 
                        exit(1)
                        
                    layerRow = str(vars()[layer + '_range'][0])
                    expandCellsValue = ws.cell(expandCellsColumn + layerRow).value

                    if expandCellsValue > 0 and task == 'RESISTANCE' and doExpandCells == True:                    
                        if method == 'MINIMUM' or method == "'MINIMUM'": 
                            gprint('    ***MINIMUM value for reclassified layer ' + layer + ' will be expanded by ' + str(expandCellsValue) + ' cell(s).')
                        else:
                            gprint('    ***Maximum value for reclassified layer ' + layer + ' will be expanded by ' + str(expandCellsValue) + ' cell(s).')

                        arcpy.env.extent = os.path.join(layerFolder,layer)
                        neighborhood = arcpy.sa.NbrCircle(str(int(expandCellsValue)), "CELL")
                        if method == 'MINIMUM' or method == "'MINIMUM'":
                            outFocalStatistics = arcpy.sa.FocalStatistics(outReclass, neighborhood, "MINIMUM","")      
                        else:
                            outFocalStatistics = arcpy.sa.FocalStatistics(outReclass, neighborhood, "MAXIMUM","")      
                        outFocalStatistics.save(os.path.join(str(scratchGDB), 'l' + str(counter) + '_hab'))  
                        arcpy.env.workspace = layerFolder                        
                        arcpy.env.extent = referenceLayer
                       
                    else:    
                        outReclass.save(os.path.join(str(scratchGDB), 'l' + str(counter) + '_hab'))                     
                        arcpy.env.workspace = layerFolder
                     
                # arcpy.env.workspace = outputGDB
                arcpy.env.workspace = scratchGDB
                
                #Divide each reclassified Raster by 1000 to get back to decimal 
                gprint('\nPOSTPROCESSING LAYERS:') 
                for i in range(1, len(usedlayers) + 1, 1):
                    gprint('     ' + usedlayers[i-1])
                    outFloat = arcpy.sa.Float('l' + str(i) + '_hab')
                    # outFloat.save('l' + str(i) + '_hab_f')
                    
                    # outDec = arcpy.sa.Divide('l' + str(i) + '_hab_f', 1000)
                    outDec = arcpy.sa.Divide(outFloat, 1000)
                    del outFloat
                    try:
                        outDec.save('l' + str(i) + '_hab_dec')
                    except:
                        gprint('failed to save.')
                        delete_data('l' + str(i) + '_hab_dec')
                        outDec.save('l' + str(i) + '_hab_dec')
                    
                    del outDec
                        
                    delete_data('l' + str(i) + '_hab')
                    # delete_data('l' + str(i) + '_hab_f')
                    
                gprint('\nCOMBINING LAYERS')
                referenceLayer = os.path.join(layerFolder,referenceLayer)
                arcpy.env.cellSize = referenceLayer
                arcpy.env.extent = referenceLayer
                arcpy.env.snapRaster = referenceLayer
                gprint('Spatial reference layer: ' + referenceLayer)
                
                if method == 'PRODUCT' or method == "'PRODUCT'":
                    # multiply all layers together and write to output folder 
                    # (map algebra)        
                    operation = []
                    for i in range(1, len(usedlayers) + 1, 1):
                        operation.append('l' + str(i) + '_hab_dec' + " *")
                    operation = ",".join(operation).replace(",", " ")
                    operation = operation[:-2]
                    if task == 'HABITAT':
                        outfilename = os.path.join(outputGDB,tableBase + "_Habitat_prod")
                    else:
                        outfilename = os.path.join(outputGDB,tableBase + "_R_prod")

                    # arcpy.env.workspace = outputGDB

                    gprint('\nMULTIPLYING LAYERS TOGETHER\n')                         
                    gp.SingleOutputMapAlgebra_sa(str(operation), outfilename)  
                    build_stats(outfilename)        

                else:
                    rasterList = []
                    for i in range(1, len(usedlayers) + 1, 1):
                        rasterList.append('l' + str(i) + '_hab_dec')
                    # arcpy.env.workspace = outputGDB
                    
                gprint(' ')    
                if method == 'SUM' or method == "'SUM'":      
                    if task == 'HABITAT':
                        outfilename = os.path.join(outputGDB,tableBase + "_Habitat_sum")
                    else:
                        outfilename = os.path.join(outputGDB,tableBase + "_R_sum")
                    
                    gprint('ADDING LAYERS TOGETHER\n')      
                    
                    addValue = 0
                    if addOneToResist:
                        addValue = 1
                    outCellStatistics = arcpy.sa.Plus(arcpy.sa.CellStatistics(rasterList, "SUM", "DATA"), addValue)
                    
                    outCellStatistics.save(outfilename)
                    build_stats(outfilename)        
                    #create_r_sens_layers(outfilename)
                    
                elif method == 'MINIMUM' or method == "'MINIMUM'": 
                    if task == 'HABITAT':
                        outfilename = os.path.join(outputGDB,tableBase + "_Habitat_min")
                    else:
                        outfilename = os.path.join(outputGDB,tableBase + "_R_min")

                    # calculate minimum value of all layers and write to output folder
                    # (map algebra)              
                    gprint('CALCULATING MINIMUM %s VALUE\n' %task)

                    # clearWSLocks(outputGDB)
                    # clearWSLocks(scratchGDB)
                    
                    outCellStatistics = arcpy.sa.CellStatistics(rasterList, "MINIMUM", "DATA")
                    outCellStatistics.save(outfilename)
                    build_stats(outfilename)        
                                       
                elif method == 'MAXIMUM' or method == "'MAXIMUM'": 
                    method = habitatMethod
                    if task == 'HABITAT':
                        outfilename = os.path.join(outputGDB,tableBase + "_Habitat_max")
                    else:
                        outfilename = os.path.join(outputGDB,tableBase + "_R_max")
                    gprint('CALCULATING MAXIMUM ' + task + ' VALUE' + '\n')
                    outCellStatistics = arcpy.sa.CellStatistics(rasterList, "MAXIMUM", "DATA")
                    outCellStatistics.save(outfilename)
                    build_stats(outfilename)        
                    #create_r_sens_layers(outfilename)   
                gprint('Output Geodatabase: %s\n' %outputGDB)                    
                gprint('Output raster name: %s\n' %outfilename)
                gprint('DELETING USED LAYERS\n')
                for i in range(1, len(usedlayers) + 1, 1):
                    delete_data('l' + str(i) + '_hab_dec')
                
                # Clean up
                arcpy.env.workspace = outputGDB
                arcpy.env.scratchWorkspace = outputGDB
                clean_out_ws(scratchGDB)

                delete_data(scratchGDB)
                delete_dir(scratchGDB)
                delete_dir(scratchDir)
        gprint('Done!')
        gprint('\n-----------------------------------------------------------------')
        gprint('If you use this software, please cite it so others can find it!')
        gprint('See www.circuitscape.org/gnarly-landscape-utilities \nfor preferred citation')
        gprint('-----------------------------------------------------------------')
        
    # Return GEOPROCESSING specific errors  
    except arcgisscripting.ExecuteError: 
        gprint('****Geoprocessing error. Details follow.****') 
        raise_geoproc_error(__filename__) 

    # Return any PYTHON or system specific errors  
    except:
        gprint('****Python error. Details follow.****') 
        raise_python_error(__filename__) 
            
    
    
# Define functions
def unique(seq):
    checked = []
    for i in seq:
        if i not in checked:
            checked.append(i)
    return checked

def raise_error(msg):
    gp.AddError(msg)
    write_log(msg)
    exit(1)

def raise_geoproc_error(filename): 
    """Handle geoprocessor errors and provide details to user"""
    tb = sys.exc_info()[2]  # get the traceback object
    # tbinfo contains the error's line number and the code
    tbinfo = traceback.format_tb(tb)[0]
    line = tbinfo.split(", ")[1]

    arcpy.AddError("Geoprocessing error on **" + line + "** of " + filename + " :")
    if not arcpy.GetMessages(2) == "":
        arcpy.AddError(arcpy.GetMessages(2))                                


    # for msg in range(0, gp.MessageCount):
        # if gp.GetSeverity(msg) == 2:
            # gp.AddReturnMessage(msg)
    exit(0)

def raise_python_error(filename): 
    """Handle python errors and provide details to user"""
    tb = sys.exc_info()[2]  # get the traceback object
    # tbinfo contains the error's line number and the code
    tbinfo = traceback.format_tb(tb)[0]
    line = tbinfo.split(", ")[1]

    err = traceback.format_exc().splitlines()[-1]

    arcpy.AddError("Python error on **" + line + "** of " + filename)
    arcpy.AddError(err)
    exit(0)

def build_stats(outfilename):
        try:
            # generate pyramids and statistics for final output
            gprint('BUILDING OUTPUT STATISTICS & PYRAMIDS' + '\n')        
            arcpy.CalculateStatistics_management(outfilename, "1", "1", "#")
            arcpy.BuildPyramids_management(outfilename)       
        except:
            pass
                                       
def gprint(string):
    arcpy.AddMessage(string)
    write_log(string) 

def write_log(string):
    try:
        logFile=open(logFilePath,'a')
    except:
        logFile=open(logFilePath,'w')
    try:
        #Sometimes int objects returned for arc failures so need str below
        logFile.write(str(string) + '\n') 
    except IOError:
        pass
    finally:
        logFile.close()
        
def delete_data(dataset):
    try:
        if arcpy.Exists(dataset):
            arcpy.Delete_management(dataset)
    except:
        pass

def clean_out_ws(ws):
    try:
        if gp.exists(ws):
            gp.workspace = ws
            gp.OverwriteOutput = True
            fcs = gp.ListFeatureClasses()
            for fc in fcs:
                fcPath = os.path.join(ws,fc)
                gp.delete_management(fcPath)

            rasters = gp.ListRasters()
            for raster in rasters:
                rasterPath = os.path.join(ws,raster)
                gp.delete_management(rasterPath)
    except:
        pass

def delete_dir(dir):
    if arcpy.Exists(dir):
        try:
            arcpy.RefreshCatalog(dir)
            shutil.rmtree(dir)
        except:
            # In case rmtree was unsuccessful due to lock on data
            try:
                arcpy.RefreshCatalog(dir)
                arcpy.Delete_management(dir)
            except:
                pass
    return


def str2bool(pstr):
    """Convert ESRI boolean string to Python boolean type"""
    return pstr == 'true'

        
def clearWSLocks(inputWS):
    """Attempts to clear locks on a workspace."""
    arcpy.env.workspace = inputWS
    if all([arcpy.Exists(inputWS), arcpy.Compact_management(inputWS), arcpy.Exists(inputWS)]):
        gprint( 'Workspace (%s) clear to continue...' % inputWS)
    else:
        gprint( '!!!!!!!! ERROR WITH WORKSPACE %s !!!!!!!!' % inputWS)
                            
def check_path(path):
    """Checks to make sure path name is not too long.

    Long path names can cause problems with ESRI grids.
    """
    if len(path) > 140:
        msg = ('ERROR: Directory "' + path +
               '" is too deep.  Please choose a shallow directory'
               '(something like "C:\PUMA").')
        raise_error(msg)

    if "-" in path or " " in path or "." in path:
        msg = ('ERROR: Output directory cannot contain spaces, dashes, or '
                'special characters.')
        raise_error(msg)
    head=path
    for i in range(1,100):
        if len(head) < 4: # We've gotten to the base of the tree
            break
        head,tail=os.path.split(head)
        if tail[0].isdigit():
            msg = ('ERROR: No directory names in output path can start with a number or '
                    'else Arc may crash. Please change name of "' + tail + '" or choose a new directory.')
            raise_error(msg)
    return
    
if __name__ == "__main__":
    habitat_model_builder()
