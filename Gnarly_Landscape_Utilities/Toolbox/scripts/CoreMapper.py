#
# CoreMapper.py for ArcGIS 10
# Developed by Andrew Shirk and Brad McRae
# ---------------------------------------------------------------------------

# Import system modules
import sys
import time
import os
import shutil
import arcgisscripting
from openpyxl.reader.excel import load_workbook
import arcpy
from arcpy import env
from arcpy.sa import *
import traceback 
import gnarly_version as ver

__version__ = ver.releaseNum

gp = arcpy.gp
arcpy.CheckOutExtension("spatial")

__filename__ = "Core_Mapper.py"

GP_NULL = '#'

tif = '.tif'
tif = ''

def str2Bool(pstr):
    """Convert ESRI boolean string to Python boolean type"""
    if 'true' in pstr.lower() or 'yes' in pstr.lower():
        return True
    else:
        return False

##################################################################
# Tool inputs
if len(sys.argv) > 1: # If called from ArcMap
    tables = sys.argv[1]

else: # Not called from ArcMap
    tables = 'C:\\CoreMapper\\TEST.xlsx'
##################################################################

# Initialize log file- write to output directory of first scenario
tables = tables.split(';')
wb = load_workbook(filename=tables[0])
ws = wb.get_active_sheet()
outputBaseFolder1 = ws.cell('D2').value
messageDir = os.path.join(outputBaseFolder1,'log')

if not os.path.exists(outputBaseFolder1):
    gp.CreateFolder_management(os.path.dirname(outputBaseFolder1),
                               os.path.basename(outputBaseFolder1))

if not os.path.exists(messageDir):
    gp.CreateFolder_management(os.path.dirname(messageDir),
                               os.path.basename(messageDir))
       
ft = tuple(time.localtime())
timeNow = time.ctime()
fileName = ('%s_%s_%s_%s%s_%s.txt' % (ft[0], ft[1], ft[2], ft[3], ft[4], 
            'Core_Mapper'))
logFilePath = os.path.join(messageDir,fileName)
try:
    logFile=open(logFilePath,'a')
except:
    logFile=open(logFilePath,'w')
if sys.argv is not None:
    logFile.write('*'*70 + '\n')
    logFile.write('Core Mapper log file: %s \n\n' % ('Core'))
    logFile.write('Start time:\t%s \n' % (timeNow))
    logFile.write('Parameters:\t%s \n\n' % (sys.argv))
logFile.close()
##################################################################

        
def core_mapper():
    try:   
        gprint('\nCore Mapper version ' + __version__)
        arcpy.env.overwriteOutput = 1
        lastScratchDir = 'None'
        
        # Read each excel file
        for table in tables:
            # Open and read input file
            gprint('\n***************************************')            
            gprint('READING ' + table + '\n')
            wb = load_workbook(filename=table)
            ws = wb.get_active_sheet()

            datarange = 'A2' + ':A' + str(len(ws.row_dimensions))
            variants = ws.range(datarange)

            for i in range(len(variants)):
                variant = i + 2
                # Get inputs for model run
                outputBaseName = ws.cell('A' + str(variant)).value                
                habitatRaster = ws.cell('B' + str(variant)).value
                resistanceRaster = nullstring(ws.cell('C' + str(variant)).value)
                outputBaseFolder = ws.cell('D' + str(variant)).value
                gprint('\n***************************************')            
                gprint('PROCESSING ' + outputBaseName + ' run.\n')

                movingWindowRadius = int(ws.cell('E' + str(variant)).value)     ## Euclidean distance
                minAvgHabValue = float(ws.cell('F' + str(variant)).value)       ## used in the moving window 
                binaryThreshold = float(ws.cell('G' + str(variant)).value)      ## used to convert habitat model to binary 

                expandCWDValue = nullfloat(int(ws.cell('H' + str(variant)).value))        ## cost-weighted distance
                if expandCWDValue > 0 and resistanceRaster is None:
                    gprint('Warning: a CWD expansion value was entered but no resistance raster was specified.')
                    gprint('Skipping expansion step.')
                    expandCWDValue = 0

                removeCWDHalos = str2Bool(ws.cell('I' + str(variant)).value)
                minCoreArea = int(ws.cell('J' + str(variant)).value)                
                stampCores = str2Bool(ws.cell('K' + str(variant)).value)      ## used to convert habitat model to binary 
                appendCoreStats = str2Bool(ws.cell('L' + str(variant)).value) 
                deleteIntermediates = str2Bool(ws.cell('M' + str(variant)).value) #xxx
                
                gprint('Starting core processing for output: ' + outputBaseName + '\n')
                
                #check for valid cutoff value
                maxObject = gp.GetRasterProperties(habitatRaster, "MAXIMUM") 
                rasterMax = float(str(maxObject.getoutput(0)))
                if binaryThreshold > rasterMax or minAvgHabValue > rasterMax:
                    gprint('***********************************************')
                    gprint('Error: cutoff value is higher than maximum value'
                            'of habitat raster.\nSkipping this iteration.')
                    gprint('***********************************************\n')
                    continue
                
                create_dir(outputBaseFolder)
                scratchDir = os.path.join(outputBaseFolder,'tmp_'+outputBaseName)
                delete_dir(scratchDir)
                create_dir(scratchDir)

                arcpy.env.workspace = scratchDir
                arcpy.env.scratchWorkspace = scratchDir
                gp.workspace = scratchDir
                gp.scratchWorkspace = scratchDir
                
                outputFolder = os.path.join(outputBaseFolder,'cores')    
                create_dir(outputFolder)
                
                gprint ('Habitat model is ' + str(habitatRaster) + '\n')   
                gprint ('Resistance model is ' + str(resistanceRaster) + '\n') 
                if expandCWDValue == 0:
                    gprint ('*************************************************************')
                    gprint ('Core areas will NOT be expanded using cost weighted distance.')
                    gprint ('*************************************************************\n')
                                   
                gprint ('Base folder is ' + outputBaseFolder + '\n')   

                # Get cell size and convert home range area to number of grid cells
                arcpy.env.cellSize = habitatRaster
                cellSize = float(arcpy.env.cellSize)
                if cellSize > movingWindowRadius: 
                    gprint('***********************************************')
                    gprint('Error: moving window radius must be larger than '
                            'cell size.\nSkipping this iteration.')
                    gprint('***********************************************\n')
                    continue
                # Calculate proportion of habitat within a movingWindowRadius-sized
                # moving window
                spatialref = arcpy.Describe(habitatRaster).spatialReference                
                mapunits = spatialref.linearUnitName
                gprint ('Calculating average habitat value within a ' + str(float(movingWindowRadius)) + ' ' + mapunits + ' moving window radius' + '\n')    
                avghabvalue = os.path.join(scratchDir,"avghabval"+str(i)+tif)

                delete_data(avghabvalue)
                gp.FocalStatistics(habitatRaster, avghabvalue, "Circle " + str(float(movingWindowRadius)) + " MAP", "MEAN", "DATA")  

                corehabitat = os.path.join(scratchDir,"corehab"+str(i)+tif)
                delete_data(corehabitat)

                gprint('Converting moving window result to binary habitat model based on moving window average >= ' + str(minAvgHabValue))
                gprint('and per-cell habitat values >= ' + str(binaryThreshold))
                gp.SingleOutputMapAlgebra_sa("Con (" + avghabvalue + " >= " + str(minAvgHabValue) + ", Con (" + habitatRaster + " >= " + str(binaryThreshold) + ", 1))",corehabitat)

                
                # For habitat cells in areas where habitat ratio is >
                # minAvgHabValue, run cost distance
                if expandCWDValue == 0:
                    gprint('\nSkipping cost-weighted distance calculations.' + '\n')
                    prelimCores = corehabitat
                    CWDCoreHab = None
                else:
                    gprint('\nExpanding cores by ' + str(float(expandCWDValue)) + ' ' + mapunits + 's in cost-weighted distance' + '\n')
                    CWDCoreHab = os.path.join(scratchDir,"CWDCoreHab"+str(i)+tif)
                    delete_data(CWDCoreHab)
                    try:
                        outCostDist = arcpy.sa.CostDistance(corehabitat, resistanceRaster, str(float(expandCWDValue)), "#")
                        outCostDist.save(CWDCoreHab)
                    except:
                        gp.AddWarning('-------------------------------------------------')
                        gp.AddWarning('Warning: cost distance mapping failed.' 
                                      '\nSkipping this iteration.')
                        gp.AddWarning('-------------------------------------------------')
                        continue

                    gprint('Done expanding cores.')
                    
                    prelimCores = os.path.join(scratchDir,"prelimCrs"+str(i)+tif)
                    delete_data(prelimCores)
                
                    outCon = arcpy.sa.Con(Raster(CWDCoreHab) <= expandCWDValue, 1)
                    outCon.save(prelimCores)
                    if removeCWDHalos and expandCWDValue > 0:
                        #Note: this can be sped up by trimming resistance raster using minAvgHabValue
                        outCon = arcpy.sa.Con(Raster(avghabvalue) > minAvgHabValue , prelimCores)
                        outCon.save(prelimCores+'2')
                        prelimCores = prelimCores + '2'
                        
                #regiongroup, using FOUR neighbors so that a diagonal (road) breaks the patch
                regionGroup = os.path.join(scratchDir,"regionGrp"+str(i)+tif)
                delete_data(regionGroup)
                try:
                    gp.RegionGroup_sa(prelimCores,regionGroup,"FOUR","WITHIN","ADD_LINK","#")
                except:
                    gprint('Region group failed.  Retrying...')
                    time.sleep(5)
                    try:
                        gp.RegionGroup_sa(prelimCores,regionGroup,"FOUR","WITHIN","ADD_LINK","#")
                    except:
                        gprint('***********************************************')
                        gprint('Region group failed.  Skipping this iteration')
                        gprint('***********************************************')
                        continue
                if stampCores and expandCWDValue > 0:
                    regionGroupStamped = regionGroup + 's'
                    outCon = arcpy.sa.Con(Raster(habitatRaster) > binaryThreshold, regionGroup)
                    outCon.save(regionGroupStamped)
                else:
                    regionGroupStamped = regionGroup

                
                zoneRaster = os.path.join(scratchDir,"zone"+str(i)+tif)
                delete_data(zoneRaster)
                zoneField = "value"
                outZonalStatistics = ZonalStatistics(regionGroup, zoneField, regionGroupStamped,
                                     "SUM", "DATA") # NOTE: If changed to below, will actually punch out low quality areas instead of just counting them against area calcs
                # outZonalStatistics = ZonalStatistics(regionGroupStamped, zoneField, regionGroupStamped,
                                     # "SUM", "DATA") 
                outZonalStatistics.save(zoneRaster)      
                areaRaster = os.path.join(scratchDir,"area"+str(i)+tif)
                outRas = Raster(zoneRaster) / Raster(regionGroup) # summing region pixels, so divide out pixel values
                outRas.save(areaRaster)
                                
                #find the patches/cores that are at least the area threshold
                MinCoreCells = str(int(0.5 + float(minCoreArea) / (cellSize *
                                    cellSize)))
                gprint('Removing cores with area less than ' + str(minCoreArea) + ' square ' + mapunits + 's')
                gprint('('+str(MinCoreCells) + ' cells).\n')

                coreRaster = os.path.join(scratchDir,"coreRaster"+str(i)+tif)
                delete_data(coreRaster) 
                outCon = arcpy.sa.Con(Raster(areaRaster) > float(MinCoreCells), regionGroup)
                outCon.save(coreRaster)


                # Convert core raster zones to polygons
                gp.AddMessage('Generating final core area polygons')
                corePolygons = os.path.join(scratchDir,"corePolygons"+str(i)+".shp")
                delete_data(corePolygons)
                try:
                    arcpy.RasterToPolygon_conversion(coreRaster, corePolygons, "NO_SIMPLIFY", "VALUE")
                except:
                    gprint('********************************************')
                    gprint('NO cores that meet critera to map!  \nSkipping this iteration.')
                    gprint('********************************************\n')
                    continue
                    
                dissolveFC = os.path.join(scratchDir, outputBaseName + "_coreDissolve.shp")
                gp.Dissolve_management (corePolygons, dissolveFC , "GRIDCODE")
                gp.AddField_management(dissolveFC, "core_ID", "LONG")
                gp.CalculateField_management(dissolveFC, "core_ID", "[FID] + 1", "VB")
                
                finalCores = os.path.join(outputFolder,outputBaseName + "_cores.shp")
                delete_data(finalCores)
                gp.CopyFeatures_management(dissolveFC,finalCores)
                
                if appendCoreStats:
                    gp.AddMessage('Appending core statistics to ' + finalCores)
                    zonalTable = os.path.join(outputFolder,outputBaseName + "_stats.dbf")
                    gp.ZonalStatisticsAsTable (finalCores, "core_ID", habitatRaster, zonalTable, "DATA", "ALL")
                    gp.JoinField_management (finalCores, "core_ID", zonalTable, "core_ID")
                    arcpy.DeleteField_management(finalCores, ["core_ID_1", "GRIDCODE"]) 
                    delete_data(zonalTable)                    
                             
                if deleteIntermediates:
                    gprint('Deleting intermediate files' + '\n')
                    delete_data(corehabitat)
                    delete_data(CWDCoreHab)
                    delete_data(prelimCores)
                    delete_data(regionGroup)
                    delete_data(regionGroupStamped)
                    # delete_data(regionArea)
                    delete_data(areaRaster)
                    delete_data(prelimCores)
                    delete_data(avghabvalue)
                    delete_data(coreRaster)
                    delete_data(dissolveFC)
                    delete_data(corePolygons)
                    delete_dir(scratchDir)                        
                    delete_dir(lastScratchDir)
                    lastScratchDir = scratchDir
                gprint('\nCore processing complete.\n')
                
    # Return GEOPROCESSING specific errors  
    except arcpy.ExecuteError: 
        gprint('****Geoprocessing error. Details follow.****') 
        raise_geoproc_error(__filename__) 
    # Return any PYTHON or system specific errors  
    except:
        gprint('****Python error. Details follow.****') 
        raise_python_error(__filename__) 
        
def gprint(string):
    gp.addmessage(string)
    write_log(string)

def create_dir(lmfolder):
    """Creates folder if it doesn't exist."""
    if not os.path.exists(lmfolder):
        gp.CreateFolder_management(os.path.dirname(lmfolder),
                                       os.path.basename(lmfolder))    

def write_log(string):
    try:
        logFile=open(logFilePath,'a')
    except:
        logFile=open(logFilePath,'w')
    try:
        #Sometimes int objects returned for arc failures so need str below
        logFile.write(str(string) + '\n') 
    except IOError:
        pass
    finally:
        logFile.close()

def copy_excel_table(table,outputdir):
    try:
        shutil.copyfile(table,os.path.join(outputdir,os.path.basename(table)))  
    except:
       pass

def nullfloat(innum):
    """Convert ESRI float or null to Python float"""
    if innum == GP_NULL:
        nfloat = None
    else:
        nfloat = float(innum)
    return nfloat


def nullstring(arg_string):
    """Convert ESRI nullstring to Python null"""
    if arg_string == GP_NULL:
        arg_string = None
    return arg_string

    
def raise_geoproc_error(filename): 
    """Handle geoprocessor errors and provide details to user"""
    tb = sys.exc_info()[2]
    tbinfo = traceback.format_tb(tb)[0]
    line = tbinfo.split(", ")[1]

    gp.AddError("Geoprocessing error on **" + line + "** of " + filename +
                " :")
    if not arcpy.GetMessages(2) == "":
            arcpy.AddError(arcpy.GetMessages(2))
        
    gprint('\n Try a new output directory.  Sometimes that does the trick.')    
    exit(0)

def raise_python_error(filename): 
    """Handle python errors and provide details to user"""
    tb = sys.exc_info()[2]
    tbinfo = traceback.format_tb(tb)[0]
    line = tbinfo.split(", ")[1]

    err = traceback.format_exc().splitlines()[-1]

    arcpy.AddError("Python error on **" + line + "** of " + filename)
    arcpy.AddError(err)
    exit(0)

def delete_data(dataset):
    try:
        if gp.Exists(dataset):
            gp.delete_management(dataset)
    except:
        pass

def delete_dir(dir):        
    if gp.Exists(dir):
        try:        
            gc.collect()
            shutil.rmtree(dir)
        except:
            pass
            
if __name__ == "__main__":
    core_mapper()
